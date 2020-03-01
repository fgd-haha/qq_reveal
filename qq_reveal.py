import ssl
import time

from openpyxl import load_workbook
from selenium import webdriver
import requests
import openpyxl

def get_qq(path):
    qq_list = {}
    wb = load_workbook(path)
    ws = wb["Sheet1"]
    ws.iter_rows()

    return ws.rows


def reveal_check(qq_rows):
    result = set()
    driver = webdriver.Chrome(r"C:\Users\fgdfr\Desktop\code\qq_reveal\config\chromedriver.exe")
    driver.maximize_window()
    driver.get('https://haveibeenpwned.com/')
    for name, email in qq_rows:
        if is_revealed(driver, email._value):
            result.add((name._value, email._value))

    return result

def is_revealed(driver, email):
    driver.find_element_by_id("Account").clear()
    driver.find_element_by_id("Account").send_keys(email)
    driver.find_element_by_id("searchPwnage").click()
    time.sleep(3)
    result = driver.find_elements_by_class_name("pwnTitle")[1].find_element_by_tag_name('h2').text

    if (result == 'Oh no — pwned!'):
        return True
    else:
        return False


def main():
    qq_rows = get_qq(r"C:\Users\fgdfr\Desktop\code\qq_reveal\data\邮箱.xlsx")
    result = reveal_check(qq_rows)
    print(result)


if __name__ == '__main__':
    main()
